import customtkinter
import tkinter as tk
from tkinter import filedialog
import os
import time
from threading import Thread, Event
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("green")

class ExcelApp(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Excel Processor with Selenium Automation")
        self.geometry("800x600")
        self.minsize(400, 300)

        self.file_path = None
        self.selected_option = "מסך החלטה(ללא אישורים)"
        self.url_map = {
            "מסך החלטה(ללא אישורים)": "https://www.selenium.dev/selenium/web/web-form.html",
            "9190 - ניהול כוח אדם וארגון": "https://example.com/9190_form",  # Поменяй на реальный URL
            "Option 3": "https://www.selenium.dev/selenium/web/web-form.html"
        }

        self.pause_event = Event()
        self.stop_event = Event()

        self.report_data = []

        self.init_main_screen()

    def init_main_screen(self):
        self.configure(fg_color="#f4f4f4")
        self.geometry("600x450")

        self.main_frame = customtkinter.CTkFrame(self, corner_radius=10, fg_color="#f4f4f4")
        self.main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        self.container = customtkinter.CTkFrame(self.main_frame, fg_color="#f4f4f4")
        self.container.pack(expand=True)

        self.upload_label = customtkinter.CTkLabel(
            self.container, text="Upload Excel File",
            font=customtkinter.CTkFont(size=16, weight="bold"),
            text_color="#222222"
        )
        self.upload_label.pack(pady=(10, 4))

        self.upload_btn = customtkinter.CTkButton(
            self.container, text="📂 Browse File", command=self.upload_file,
            width=220, height=48,
            font=customtkinter.CTkFont(size=15),
            fg_color="#3b82f6", hover_color="#2563eb"
        )
        self.upload_btn.pack(pady=2)

        self.file_name_label = customtkinter.CTkLabel(
            self.container, text="", font=customtkinter.CTkFont(size=13),
            text_color="#555555"
        )
        self.file_name_label.pack(pady=(4, 12))

        self.option_var = tk.StringVar(value="מסך החלטה(ללא אישורים)")
        self.option_menu = customtkinter.CTkComboBox(
            self.container,
            values=list(self.url_map.keys()),
            variable=self.option_var,
            width=220,
            font=customtkinter.CTkFont(size=14),
            command=self.option_selected
        )
        self.option_menu.pack(pady=(0, 6))

        self.selected_label = customtkinter.CTkLabel(
            self.container, text="Selected: מסך החלטה(ללא אישורים)",
            font=customtkinter.CTkFont(size=13),
            text_color="#444444"
        )
        self.selected_label.pack(pady=(0, 16))

        self.run_btn = customtkinter.CTkButton(
            self.container, text="🚀 Start Automation", command=self.start_processing,
            width=250, height=52,
            font=customtkinter.CTkFont(size=16, weight="bold"),
            fg_color="#10b981", hover_color="#059669"
        )
        self.run_btn.pack()

        self.progress = customtkinter.CTkProgressBar(self.container, height=16)
        self.progress.set(0)

        self.time_label = customtkinter.CTkLabel(
            self.container, text="Estimated time left: --",
            font=customtkinter.CTkFont(size=13),
            text_color="#333333"
        )

        self.controls = customtkinter.CTkFrame(self.container, fg_color="#f4f4f4")

        self.pause_btn = customtkinter.CTkButton(
            self.controls, text="Pause", command=lambda: self.pause_event.clear(),
            fg_color="#facc15", hover_color="#eab308"
        )
        self.continue_btn = customtkinter.CTkButton(
            self.controls, text="Continue", command=lambda: self.pause_event.set(),
            fg_color="#4ade80", hover_color="#22c55e"
        )
        self.stop_btn = customtkinter.CTkButton(
            self.controls, text="Stop", command=lambda: self.stop_event.set(),
            fg_color="#f87171", hover_color="#ef4444"
        )

        self.pause_btn.pack(side="left", padx=8)
        self.continue_btn.pack(side="left", padx=8)
        self.stop_btn.pack(side="left", padx=8)

    def upload_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file_path = path
            self.file_name_label.configure(text=os.path.basename(path), text_color="black")

    def option_selected(self, choice=None):
        self.selected_option = self.option_var.get()
        self.selected_label.configure(text=f"Selected: {self.selected_option}")

    def start_processing(self):
        if not self.file_path:
            self.file_name_label.configure(text="Please upload an Excel file first!", text_color="red")
            return

        try:
            self.wb = load_workbook(self.file_path)
            self.ws = self.wb.active
            # В зависимости от опции заголовки разные
            if self.selected_option == "9190 - ניהול כוח אדם וארגון":
                required_columns = ["Mispar Ishi", "Sug Minui", "Date Start", "Number"]
            else:
                required_columns = ["Start Date", "End Date", "Number", "Makom", "Darga", "Hail", "Isug"]

            for col in required_columns:
                if col not in [cell.value for cell in self.ws[1]]:
                    raise ValueError(f"Missing column: {col}")

            headers = [cell.value for cell in self.ws[1]]
            self.col_indexes = {col: headers.index(col) for col in required_columns}
            self.total_rows = self.ws.max_row - 1

        except Exception as e:
            self.file_name_label.configure(text=f"Error reading file: {e}", text_color="red")
            return

        self.pause_event.set()
        self.stop_event.clear()
        self.report_data = []

        self.upload_label.pack_forget()
        self.upload_btn.pack_forget()
        self.file_name_label.pack_forget()
        self.option_menu.pack_forget()
        self.selected_label.pack_forget()
        self.run_btn.pack_forget()

        self.progress.pack(pady=(10, 5), padx=40, fill="x")
        self.progress.set(0)
        self.time_label.pack(pady=(0, 20))
        self.controls.pack(pady=20)

        Thread(target=self.automation_task, daemon=True).start()

    def format_time(self, seconds):
        minutes = int(seconds) // 60
        seconds = int(seconds) % 60
        return f"{minutes} min {seconds} sec" if minutes > 0 else f"{seconds} sec"

    def automation_task(self):
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service)
        except Exception as e:
            self.time_label.configure(text=f"WebDriver error: {e}")
            return

        start_time = time.time()
        url = self.url_map.get(self.selected_option, "https://www.selenium.dev/selenium/web/web-form.html")

        for i, row in enumerate(self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, values_only=True)):
            data = {key: str(row[self.col_indexes[key]]).strip() if row[self.col_indexes[key]] else "" for key in self.col_indexes}

            report_row = data.copy()

            if self.stop_event.is_set():
                report_row["Status"] = "Processing stopped by user"
                self.report_data.append(report_row)
                break

            self.pause_event.wait()

            try:
                driver.get(url)
                time.sleep(0.5)

                if self.selected_option == "9190 - ניהול כוח אדם וארגון":
                    # Новый алгоритм для второй опции

                    # Вставить миспар иши, нажать Enter
                    mispar_ishi_input = driver.find_element(By.XPATH, "//input[@name='mispar_ishi']")
                    #mispar_ishi_input.clear()
                    mispar_ishi_input.send_keys(data["Mispar Ishi"])
                    mispar_ishi_input.send_keys(Keys.ENTER)
                    time.sleep(0.5)

                    # Вставить номер 9190, нажать Enter
                    number_9190_input = driver.find_element(By.XPATH, "//input[@name='number_9190']")
                    #number_9190_input.clear()
                    number_9190_input.send_keys("9190")
                    number_9190_input.send_keys(Keys.ENTER)
                    time.sleep(0.5)

                    # Открыть дропдаун, выбрать первый элемент
                    dropdown = driver.find_element(By.XPATH, "//select[@name='dropdown']")
                    dropdown.click()
                    time.sleep(0.5)
                    first_option = driver.find_element(By.XPATH, "//select[@name='dropdown']/option[2]")
                    first_option.click()
                    time.sleep(0.5)

                    # Ожидаем загрузку нового окна/формы (примитивный sleep, можно через WebDriverWait)
                    time.sleep(2)

                    # Заполняем новую форму
                    # 1) миспар иши
                    new_mispar_ishi = driver.find_element(By.XPATH, "//input[@name='new_mispar_ishi']")
                    new_mispar_ishi.clear()
                    new_mispar_ishi.send_keys(data["Mispar Ishi"])

                    # 2) суг минуй
                    sug_minui = driver.find_element(By.XPATH, "//input[@name='sug_minui']")
                    sug_minui.clear()
                    sug_minui.send_keys(data["Sug Minui"])

                    # 3) дата старт
                    date_start = driver.find_element(By.XPATH, "//input[@name='date_start']")
                    date_start.clear()
                    date_start.send_keys(data["Date Start"])

                    # 4) номера
                    number_field = driver.find_element(By.XPATH, "//input[@name='number']")
                    number_field.clear()
                    number_field.send_keys(data["Number"])

                    number_field.send_keys(Keys.ENTER)
                    time.sleep(10)

                    # Нажать кнопку скачать (пример xpath)
                    download_btn = driver.find_element(By.XPATH, "//button[contains(text(),'Download')]")
                    download_btn.click()

                    report_row["Status"] = "Sucsess"

                else:
                    # Текущая логика для первой и третьей опций
                    driver.find_element(By.XPATH, "//input[@type='checkbox']").click()
                    time.sleep(0.5)

                    driver.find_element(By.NAME, "my-text").clear()
                    driver.find_element(By.NAME, "my-text").send_keys(data.get("Number", ""))

                    driver.find_element(By.NAME, "my-password").clear()
                    driver.find_element(By.NAME, "my-password").send_keys(data.get("Makom", ""))

                    driver.find_element(By.NAME, "my-textarea").clear()
                    driver.find_element(By.NAME, "my-textarea").send_keys(data.get("Darga", ""))

                    driver.find_element(By.NAME, "my-datalist").clear()
                    driver.find_element(By.NAME, "my-datalist").send_keys(data.get("Hail", ""))

                    driver.find_element(By.XPATH, "//input[@type='radio']").click()
                    time.sleep(0.5)

                    result_text = driver.find_element(By.TAG_NAME, "h1").text.strip()
                    report_row["Status"] = result_text

            except Exception as e:
                report_row["Status"] = f"Error: {str(e)}"

            self.report_data.append(report_row)

            self.progress.set((i + 1) / self.total_rows)
            elapsed = time.time() - start_time
            avg_time = elapsed / (i + 1)
            remaining_time = avg_time * (self.total_rows - (i + 1))
            self.time_label.configure(text=f"Estimated time left: {self.format_time(remaining_time)}")
            time.sleep(0.5)

        driver.quit()
        self.progress.pack_forget()
        self.controls.pack_forget()
        self.time_label.pack_forget()

        finish_message = "✅ Completed successfully" if not self.stop_event.is_set() else "⛔ Processing stopped by user"

        self.finish_label = customtkinter.CTkLabel(
            self.container, text=finish_message,
            font=customtkinter.CTkFont(size=16, weight="bold"),
            text_color="#10b981" if not self.stop_event.is_set() else "#dc2626"
        )
        self.finish_label.pack(pady=(20, 12))

        self.download_btn = customtkinter.CTkButton(
            self.container, text="💾 Download Report", command=self.download_report,
            width=240, height=48,
            font=customtkinter.CTkFont(size=15, weight="bold"),
            fg_color="#3b82f6", hover_color="#2563eb"
        )
        self.download_btn.pack(pady=(0, 10))

        self.return_btn = customtkinter.CTkButton(
            self.container, text="🔁 Return to Main", command=self.reset_ui,
            width=240, height=48,
            font=customtkinter.CTkFont(size=15, weight="bold"),
            fg_color="#a855f7", hover_color="#9333ea"
        )
        self.return_btn.pack()

    def download_report(self):
        file_save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            title="Save Report",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_save_path:
            try:
                wb_report = Workbook()
                ws_report = wb_report.active
                # В зависимости от опции разный заголовок отчёта
                if self.selected_option == "9190 - ניהול כוח אדם וארגון":
                    headers = ["Mispar Ishi", "Sug Minui", "Date Start", "Number", "Status"]
                else:
                    headers = ["Start Date", "End Date", "Number", "Makom", "Darga", "Hail", "Isug", "Status"]
                ws_report.append(headers)
                for row in self.report_data:
                    ws_report.append([row.get(col, "") for col in headers])
                wb_report.save(file_save_path)
                self.time_label.configure(text=f"Report saved: {os.path.basename(file_save_path)}")
            except Exception as e:
                self.time_label.configure(text=f"Error saving report: {e}")

    def reset_ui(self):
        if hasattr(self, 'download_btn'):
            self.download_btn.pack_forget()
        if hasattr(self, 'return_btn'):
            self.return_btn.pack_forget()
        if hasattr(self, 'finish_label'):
            self.finish_label.pack_forget()

        self.time_label.configure(text="")
        self.file_name_label.configure(text="")

        self.upload_label.pack(pady=(10, 4))
        self.upload_btn.pack(pady=2)
        self.file_name_label.pack(pady=(4, 12))
        self.option_menu.pack(pady=(0, 6))
        self.selected_label.pack(pady=(0, 16))
        self.run_btn.pack()

        self.file_path = None


if __name__ == "__main__":
    app = ExcelApp()
    app.mainloop()
